import satnogs_api
import openpyxl  # excel
from datetime import datetime


DIR = 'D:\\satnogs_satellites.xlsx'


def satelliteFilter(satMap: dict, mode="AFSK", baud=1200, time=365) -> dict:
    pass


def sortRecent(satMap: dict) -> dict:  # sort by most recently updated
    pass


def satelliteMap2Excel(satMap: dict) -> None:
    if not satMap:  # if map is empty
        return

    wb = openpyxl.Workbook()  # create a openpyxl Workbook object
    ws = wb.create_sheet("Unfiltered", 0)  # create tab for original JSON output

    ws["A1"] = "Name"
    ws["B1"] = "Description"
    ws["C1"] = "Norad_cat_id"
    ws["D1"] = "Service"
    ws["E1"] = "Mode"
    ws["F1"] = "Baud Rate"
    ws["G1"] = "Timestamp"

    currRow = 2
    for entry in satMap.values():
        ws.cell(currRow, 1, entry["name"])
        ws.cell(currRow, 2, entry["description"])
        ws.cell(currRow, 3, entry["norad_cat_id"])
        ws.cell(currRow, 4, entry["service"])
        ws.cell(currRow, 5, entry["mode"])
        ws.cell(currRow, 6, entry["baud"])
        ws.cell(currRow, 7, entry["time"])
        currRow += 1

    ws = wb.create_sheet("Filtered", 1)  # create tab for filtered output
    
    # code to output filtered result
    
    ws = wb.create_sheet("Sorted", 1)  # create tab for sorted output
    
    # code to output most recent result

    # save Workbook to Excel file
    wb.save(DIR)

    return


# driver
satelliteDict = satnogs_api.getSatellites()
for v in satelliteDict.values():
    print(v, end="\n")

satelliteMap2Excel(satelliteDict)